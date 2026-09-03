"""התאמת ספרים מול דפי בנק — התאמה כפולה.

דפי הבנק של החודש מועלים ידנית לתת-תיקייה חסויה בתוך תיקיית החודש:
    <בסיס>/YYYY-MM חודש/_בנק/

עקרון החסיון: דפי הבנק לעולם אינם מצורפים לאף מייל ואינם נסרקים
כחשבוניות. רק תוצאת ההתאמה (שורות תואמות/חסרות) נכנסת לדוח.

ההתאמה הכפולה:
    כיוון א: כל חשבונית ↔ האם קיימת תנועת בנק בסכום שלה (שולם/התקבל?)
    כיוון ב: כל תנועת בנק ↔ האם קיים לה מסמך (תנועה ללא חשבונית = חסר תיעוד)

נתמכים: CSV ו-XLSX (ייצוא מאתר הבנק, כל בנק ישראלי), וגם PDF של דף
חשבון (חילוץ סכומים בלבד). ההתאמה לפי סכום בסובלנות שקל, שיטה שמרנית:
עדיף התאמה חסרה שתיבדק בעין מאשר התאמת שווא.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

from .accounting import BANK_SUBFOLDER, Row, extract_text

AMOUNT_TOLERANCE = 1.00

_AMT = re.compile(r"-?\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?|-?\d+(?:\.\d{1,2})?")


@dataclass
class BankTx:
    """תנועת בנק אחת."""
    amount: float          # חיובי = זכות (כניסה), שלילי = חובה (יציאה)
    description: str = ""
    date: str = ""
    source_file: str = ""
    matched: bool = False


@dataclass
class Reconciliation:
    """תוצאת ההתאמה הכפולה."""
    statements_found: int = 0
    transactions: int = 0
    matched: list = field(default_factory=list)        # (Row, BankTx)
    unmatched_rows: list = field(default_factory=list)  # חשבוניות בלי תנועה
    unmatched_txs: list = field(default_factory=list)   # תנועות בלי חשבונית


_DATE_LIKE = re.compile(r"^\d{1,4}[./-]\d{1,2}[./-]\d{1,4}$")


def _parse_amount(s: str) -> float | None:
    s = str(s).strip().replace("₪", "").replace('"', "")
    if _DATE_LIKE.match(s):
        return None
    neg = s.startswith("-") or s.endswith("-") or ("(" in s and ")" in s)
    m = _AMT.search(s.replace("(", "").replace(")", ""))
    if not m:
        return None
    try:
        v = abs(float(m.group().replace(",", "")))
    except ValueError:
        return None
    if v == 0:
        return None
    return -v if neg else v


# זיהוי עמודות לפי שורת הכותרת של ייצוא בנק ישראלי (לאומי ודומיו)
_HDR_DEBIT = ("חובה", "חיוב")
_HDR_CREDIT = ("זכות", "זיכוי")
_HDR_DESC = ("תיאור", "פרטים", "אסמכתא נגדית")
_HDR_DATE = ("תאריך",)
_HDR_SKIP = ("יתרה", "אסמכתא", "סימוכין")


def _parse_csv_with_header(rows: list[list[str]], header: list[str], source: str) -> list[BankTx]:
    """פרסינג לפי כותרות: חובה = יציאה (שלילי), זכות = כניסה (חיובי)."""
    def col(names) -> int | None:
        for i, h in enumerate(header):
            hs = h.strip()
            if any(n in hs for n in names) and not any(s in hs for s in _HDR_SKIP if s not in names):
                return i
        return None

    debit_i, credit_i = col(_HDR_DEBIT), col(_HDR_CREDIT)
    desc_i = col(_HDR_DESC)
    date_i = col(_HDR_DATE)
    txs = []
    for cells in rows:
        if len(cells) <= max(x for x in (debit_i, credit_i) if x is not None):
            continue
        debit = _parse_amount(cells[debit_i]) if debit_i is not None and cells[debit_i].strip() else None
        credit = _parse_amount(cells[credit_i]) if credit_i is not None and cells[credit_i].strip() else None
        amount = credit if credit else (-abs(debit) if debit else None)
        if amount is None:
            continue
        txs.append(BankTx(
            amount=amount,
            description=(cells[desc_i].strip() if desc_i is not None and desc_i < len(cells) else "")[:60],
            date=(cells[date_i].strip() if date_i is not None and date_i < len(cells) else ""),
            source_file=source,
        ))
    return txs


def _parse_csv(path: Path) -> list[BankTx]:
    txs = []
    for enc in ("utf-8-sig", "cp1255", "utf-8"):
        try:
            lines = path.read_text(encoding=enc).splitlines()
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        return txs

    parsed = list(csv.reader(lines))
    # מסלול מדויק: שורת כותרת עם עמודות חובה/זכות (הייצוא הישראלי המקובל).
    # בלעדיה, נפילה לחיפוש ההיוריסטי הישן.
    for idx, row in enumerate(parsed[:5]):
        joined = ",".join(row)
        if any(h in joined for h in _HDR_DEBIT) and any(h in joined for h in _HDR_CREDIT):
            return _parse_csv_with_header(parsed[idx + 1:], row, path.name)

    for row in parsed:
        cells = [c.strip() for c in row if c.strip()]
        if len(cells) < 2:
            continue
        # ייצוא בנק לא מצוטט: "11,800.00" מתפצל ל-"11" ו-"800.00". מאחים.
        merged = []
        i = 0
        while i < len(cells):
            if (i + 1 < len(cells)
                    and re.fullmatch(r"-?\d{1,3}", cells[i])
                    and re.fullmatch(r"\d{3}(?:\.\d{1,2})?", cells[i + 1])):
                merged.append(cells[i] + cells[i + 1])
                i += 2
            else:
                merged.append(cells[i])
                i += 1
        cells = merged
        # מזהים את עמודות הסכום: חובה/זכות נפרדות או עמודה אחת עם סימן
        amounts = [(_parse_amount(c), i) for i, c in enumerate(cells)]
        amounts = [(v, i) for v, i in amounts if v is not None and abs(v) >= 1]
        if not amounts:
            continue
        # העמודה הכספית האחרונה בשורה היא בדרך כלל היתרה, לא התנועה,
        # כשיש יותר מסכום אחד. לוקחים את הראשון שאינו תאריך.
        value, idx = amounts[0]
        desc = " ".join(c for i, c in enumerate(cells) if i != idx)[:60]
        date_m = re.search(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", " ".join(cells))
        txs.append(BankTx(amount=value, description=desc,
                          date=date_m.group() if date_m else "",
                          source_file=path.name))
    return txs


def _parse_xlsx(path: Path) -> list[BankTx]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    txs = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            if len(cells) < 2:
                continue
            nums = [c for c in cells if isinstance(c, (int, float)) and abs(c) >= 1]
            if not nums:
                continue
            value = float(nums[0])
            desc = " ".join(str(c) for c in cells if isinstance(c, str))[:60]
            txs.append(BankTx(amount=value, description=desc, source_file=path.name))
    wb.close()
    return txs


def _parse_pdf(path: Path) -> list[BankTx]:
    """PDF של דף חשבון: חילוץ סכומים בלבד, בלי כיוון (חובה/זכות לא ידוע)."""
    text = extract_text(path)
    txs = []
    for m in _AMT.finditer(text):
        v = _parse_amount(m.group())
        if v is not None and abs(v) >= 1:
            txs.append(BankTx(amount=v, source_file=path.name,
                              description="(מתוך PDF, כיוון לא ידוע)"))
    return txs


def load_bank_transactions(month_folder: Path) -> tuple[list[BankTx], int]:
    """טוען את כל התנועות מקבצי _בנק. מחזיר (תנועות, מספר קבצים)."""
    bank_dir = month_folder / BANK_SUBFOLDER
    if not bank_dir.is_dir():
        return [], 0
    txs: list[BankTx] = []
    files = 0
    for p in sorted(bank_dir.iterdir()):
        if not p.is_file():
            continue
        suffix = p.suffix.lower()
        if suffix == ".csv":
            new = _parse_csv(p)
        elif suffix in (".xlsx", ".xls"):
            new = _parse_xlsx(p)
        elif suffix == ".pdf":
            new = _parse_pdf(p)
        else:
            continue
        if new:
            files += 1
            txs.extend(new)
    return txs, files


def reconcile(rows: list[Row], month_folder: Path) -> Reconciliation:
    """
    ההתאמה הכפולה. משווה כל שורת מסמך (בש"ח) מול תנועות הבנק לפי סכום.
    תנועה שהותאמה לא תותאם שוב (מניעת התאמה כפולה של אותה תנועה).
    """
    txs, files = load_bank_transactions(month_folder)
    rec = Reconciliation(statements_found=files, transactions=len(txs))
    if not txs:
        return rec

    # חשבון עסקה אינו אירוע תשלום ולכן אין לו תנועת בנק מקבילה. בלי
    # הסינון הזה כל פרופורמה נספרת כ"מסמך ללא תנועה" ומרעישה את הדוח.
    skip = {"proforma_in", "proforma_out"}
    for r in rows:
        if r.currency != "ILS" or r.total <= 0 or r.category in skip:
            continue
        hit = None
        for tx in txs:
            if tx.matched:
                continue
            if abs(abs(tx.amount) - r.total) <= AMOUNT_TOLERANCE:
                hit = tx
                break
        if hit:
            hit.matched = True
            rec.matched.append((r, hit))
        else:
            rec.unmatched_rows.append(r)

    rec.unmatched_txs = [t for t in txs if not t.matched]
    return rec


def build_reconciliation_block(rec: Reconciliation) -> str:
    """בלוק הטקסט לדוח. דפי הבנק עצמם לא נחשפים, רק תוצאת ההתאמה."""
    if rec.statements_found == 0:
        return ("התאמת בנק: לא הועלו דפי בנק לתיקיית _בנק החודש. "
                "להתאמה מלאה, העלה את דף החשבון (CSV/XLSX/PDF) לתת-התיקייה _בנק.")
    lines = [
        f"התאמת בנק ({rec.statements_found} קבצים, {rec.transactions} תנועות):",
        f"  הותאמו: {len(rec.matched)} מסמכים מול תנועות בנק",
    ]
    if rec.unmatched_rows:
        lines.append(f"  🟡 מסמכים ללא תנועת בנק תואמת ({len(rec.unmatched_rows)}), ייתכן שטרם שולמו/נגבו:")
        for r in rec.unmatched_rows[:15]:
            lines.append(f"     • {r.file} — {r.total:,.2f} ש\"ח")
    if rec.unmatched_txs:
        shown = [t for t in rec.unmatched_txs if abs(t.amount) >= 20][:15]
        lines.append(f"  🟡 תנועות בנק ללא מסמך ({len(rec.unmatched_txs)}), ייתכן חסר תיעוד:")
        for t in shown:
            lines.append(f"     • {abs(t.amount):,.2f} ש\"ח {t.description[:40]} ({t.date})")
    if not rec.unmatched_rows and not rec.unmatched_txs:
        lines.append("  ✅ התאמה מלאה: כל המסמכים וכל התנועות מוסברים")
    lines.append("  הערה: התאמה לפי סכום בלבד, אומדן ראשוני. דפי הבנק נשארים בתיקייה ואינם נשלחים.")
    return "\n".join(lines)
